import sys
import openpyxl
import datetime as dt

book = ""
sheet = ""
name = ""
def setup():
    """Here, the user will be prompted to enter name. Future implementation will have this function as a login system, so that you can access individual Excel worksheets. 
    """
    global book
    book = openpyxl.load_workbook('Stock_Advisor.xlsx')
    global sheet
    sheet = book.worksheets[0]
    global name
    name = raw_input("Hello! Please enter your name:  ").strip()
    validate_name(name, book)


def main_menu(name):
    choice = raw_input("Hello {}. Please enter B to enter in a new stock buy, S to enter in a new stock sell, V to view your trade history, or Q to quit: ".format(name)).strip()
   
    if choice.upper() == "B":
        enter_new_buy(sheet)
    elif choice.upper() == "S":
        enter_new_sell(sheet)
    elif choice.upper() == "V":
        view_trades()
    elif choice.upper() == "Q":
        sys.exit()


def validate_name(name,book):
    """
    This method takes in the user's name, and the Excel workbook. It will get the associated workbook.
    """
    try:
        sheet = book.get_sheet_by_name(name)
        main_menu(name)
    except KeyError:
        print("Sorry. You do not seem to have a worksheet set up currently with that name.\n")
        setup()


def enter_new_buy(sheet):
    """ 
    Takes in the user's sheet as initial input. Follows with a series of prompts about the stock info and exit strategy, then will write 
    to spreadsheet.
    """
    stock = raw_input("What is the name of the stock?: ").strip()
    price_per_share = float(raw_input("How much is the price of each share?: ").strip())
    amount_of_shares = int(raw_input("How many shares are being purchased? ").strip())
    date_bought = dt.datetime.today().strftime("%m/%d/%Y %H:%M")
    gain_goal = raw_input(
"IMPORTANT: \n\n Your exit strategy is vital to being an efficient investor. What is your target price per share in which you will exit to secure profits?:  ").strip()
    loss_goal = raw_input("What is the price per share in which you will exit to cut losses?: ").strip()
    total_cost = price_per_share * amount_of_shares
    
    confirm = raw_input("CONFIRM WITH Y OR N: \n\n You have purchased {} share(s) of {}, at a price of {} for each share, for a total of {}. Your target price to secure profits is {}, and your stop loss price in which you shall exit to cut losses is {}. Is this correct?: ".format(amount_of_shares, stock, price_per_share, total_cost, gain_goal, loss_goal)).strip()
    
    if confirm.upper() == 'N':
        print("Ok, we will return to main menu. ")
        main_menu(name)
    elif confirm.upper() == 'Y':
        print("Writing to spreadsheet.")
    
        rowValue = 1
        while sheet['A{}'.format(rowValue)].value != None:
            rowValue+= 1
        sheet.cell(row = rowValue, column = 1).value = stock
        sheet.cell(row = rowValue, column = 2).value = "${}".format(price_per_share)
        sheet.cell(row = rowValue, column = 3).value = amount_of_shares
        sheet.cell(row = rowValue, column = 4).value = date_bought
        sheet.cell(row = rowValue, column = 5).value = "${}".format(gain_goal)
        sheet.cell(row = rowValue, column = 6).value = "${}".format(loss_goal)
        sheet.cell(row = rowValue, column = 7).value = "${}".format(total_cost)
        book.save("Stock_Advisor.xlsx")
        print("Your trade has now been recorded.\n You will now be redirected to the main menu.\n")
        main_menu(name)
        

def enter_new_sell(sheet):
    """ 
    Takes in the user's sheet as initial input. Follows with a series of prompts to 
    located the exact purchase made before, to write the sale to the spreadsheet. 
    This version will assume the buyer wants to sell all the shares purchased.
    """
    stock = raw_input("Which stock are you selling:").strip()
    if find_stock_row(sheet, stock) != False:
        rowValue = find_stock_row(sheet, stock)
        print rowValue
        print "Your stock is available."
        price_per_share = float(raw_input("What price per share are you selling at: ").strip())
        amount_sold = int(raw_input("How many shares are you selling: ").strip())
        date_sold = dt.datetime.today().strftime("%m/%d/%Y %H:%M")
        total_revenue = price_per_share * amount_sold
        confirm = raw_input("CONFIRM WITH Y OR N: \n\n You are now selling {} share(s) of {}, at a price of {} for each share, for a total of {}. Is this correct?: ".format(amount_sold, stock, price_per_share, total_revenue))
        if confirm.upper() == 'N':
            print("Ok, we will return to main menu. ")
            main_menu(name)
        elif confirm.upper() == 'Y':
            print("Writing to spreadsheet.")
            total_cost = sheet.cell(row = 2, column = 7).value
            print total_cost
            net_gain = total_revenue - float(total_cost.replace("$", ""))
            sheet.cell(row = rowValue, column = 8).value = "${}".format(price_per_share)
            sheet.cell(row = rowValue, column = 9).value = amount_sold
            sheet.cell(row = rowValue, column = 10).value = date_sold
            sheet.cell(row = rowValue, column = 11).value = "${}".format(total_revenue)
            sheet.cell(row = rowValue, column = 12).value = "${}".format(net_gain)

            book.save("Stock_Advisor.xlsx")
            print("Your sale has now been recorded.\n You will now be redirected to the main menu.\n")
            main_menu(name)

    elif find_stock_row(sheet, stock) == False:
        print("Sorry, that name is not present in your records. You will be redirected to main menu")
        main_menu(name)

def find_stock_row(sheet, stock):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == stock:
                return cell.row
    return False

def view_trades(sheet):
    pass

if __name__ == "__main__":
    setup()
